"""Parse Indonesian bank transfer receipts (BCA e-banking format) into
structured transactions ready to be booked as journal entries.

Accepts one or more receipts pasted as plain text, e.g.:

    Tanggal             : 11/06/2026
    Jam                 : 10:14:01
    Jenis Transaksi     : TRANSFER KE BCA VIRTUAL ACCOUNT
    Dari Rekening       : 455-0068261
    No. BCA Virtual Account : 12608-8290369088
    Nama                : aXXXXXXXXXa
    Nama Perusahaan/Product : SHOPEE
    Jumlah Transfer     : Rp 1,970,100.00
    Total               : Rp 1,970,100.00
    Jenis Transfer      : Sekarang
    No Referensi        : 26061104327247
    Status              : Berhasil

Separators may be tabs, spaces, or colons; a new "Tanggal" line starts a new
receipt. Both 1,970,100.00 (US) and 1.970.100,00 (Indonesian) number formats
are recognised.
"""
import csv
import hashlib
import io
import re
from datetime import datetime

# label (lowercase) -> field name; longest labels matched first so that
# "Nama Perusahaan/Product" wins over "Nama".
LABELS = {
    "tanggal": "date",
    "jam": "time",
    "jenis transaksi": "tx_type",
    "dari rekening": "from_account",
    "no. bca virtual account": "va_number",
    "no bca virtual account": "va_number",
    "nama perusahaan/product": "merchant",
    "nama perusahaan / product": "merchant",
    "nama perusahaan": "merchant",
    "nama": "name",
    "jumlah transfer": "amount",
    "jumlah": "amount",
    "nominal transfer": "amount",
    "nominal transaksi": "amount",
    "nominal": "amount",
    "total": "total",
    "jenis transfer": "transfer_type",
    "no referensi": "reference",
    "no. referensi": "reference",
    "nomor referensi": "reference",
    "status": "status",
    "berita": "note",
    "keterangan": "note",
}
_SORTED_LABELS = sorted(LABELS, key=len, reverse=True)

SUCCESS_WORDS = ("berhasil", "sukses", "success", "successful")


def parse_amount(s):
    """'Rp 1,970,100.00' / 'Rp1.970.100,00' / '1970100' -> float."""
    s = re.sub(r"(?i)rp\.?", "", str(s)).strip()
    s = re.sub(r"[^\d.,-]", "", s)
    if not s:
        return 0.0
    if "," in s and "." in s:
        # last separator is the decimal mark
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        # single comma: decimal if followed by exactly 2 digits, else thousands
        head, _, tail = s.rpartition(",")
        s = head.replace(",", "") + ("." + tail if len(tail) == 2 else tail)
    elif "." in s:
        head, _, tail = s.rpartition(".")
        if len(tail) == 3:  # 1.970.100 style thousands
            s = s.replace(".", "")
    try:
        return round(float(s), 2)
    except ValueError:
        return 0.0


def parse_date(s):
    """DD/MM/YYYY, DD-MM-YYYY or YYYY-MM-DD -> ISO date string or None."""
    s = str(s).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _parse_line(line):
    low = line.lower()
    for label in _SORTED_LABELS:
        if low.startswith(label):
            rest = line[len(label):]
            # require a separator (colon/tab/2+ spaces) right after the label
            m = re.match(r"^\s*[:\t]\s*(.*)$|^\s{2,}(.*)$", rest)
            if m:
                value = (m.group(1) if m.group(1) is not None else m.group(2)) or ""
                return LABELS[label], value.strip()
    return None, None


def parse_bca_text(text):
    """Returns (transactions, warnings). Each transaction dict contains:
    date, time, tx_type, from_account, va_number, name, merchant, amount,
    transfer_type, reference, status, ok (status is success), description.
    """
    records, warnings = [], []
    current = {}

    def flush():
        if not current:
            return
        amount = current.get("amount") or current.get("total") or 0
        if not current.get("date") and not amount:
            return
        rec = {
            "date": current.get("date"),
            "time": current.get("time", ""),
            "tx_type": current.get("tx_type", ""),
            "from_account": current.get("from_account", ""),
            "va_number": current.get("va_number", ""),
            "name": current.get("name", ""),
            "merchant": current.get("merchant", ""),
            "amount": amount,
            "transfer_type": current.get("transfer_type", ""),
            "reference": current.get("reference", ""),
            "status": current.get("status", ""),
            "note": current.get("note", ""),
        }
        rec["ok"] = (not rec["status"]) or any(w in rec["status"].lower() for w in SUCCESS_WORDS)
        desc_bits = [b for b in (rec["tx_type"], rec["merchant"], rec["name"]) if b]
        rec["description"] = " — ".join(desc_bits) if desc_bits else "Bank transfer"
        if not rec["date"]:
            warnings.append("Receipt ref %s: missing/invalid date (Tanggal)" % (rec["reference"] or "?"))
        if not rec["amount"]:
            warnings.append("Receipt ref %s: missing amount (Jumlah Transfer)" % (rec["reference"] or "?"))
        records.append(rec)

    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        key, value = _parse_line(line)
        if key is None:
            continue
        if key == "date":
            if current.get("date") or current.get("amount"):
                flush()
                current = {}
            current["date"] = parse_date(value)
        elif key == "amount":
            current["amount"] = parse_amount(value)
        elif key == "total":
            current["total"] = parse_amount(value)
        else:
            # don't let a later duplicate label overwrite (e.g. two "Nama" variants)
            if key not in current or not current[key]:
                current[key] = value
    flush()
    return records, warnings


# ---------------------------------------------------------------------------
# BCA "Mutasi Rekening" CSV export (Informasi Rekening - Mutasi Rekening)
# ---------------------------------------------------------------------------
# Layout: preamble lines ("No. rekening : …", "Nama : …", "Periode : …"),
# then a header row: Tanggal Transaksi, Keterangan, Cabang, Jumlah, Saldo.
# Jumlah ends with CR (money in) or DB (money out).

def _decode(data):
    if isinstance(data, str):
        return data
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("latin-1", errors="replace")


def parse_bca_csv(data):
    """Parse a BCA mutasi-rekening CSV export.

    Returns (records, warnings, meta). Records share the shape of
    parse_bca_text() rows plus: direction ('in'|'out') and balance.
    """
    text = _decode(data)
    records, warnings, meta = [], [], {}
    in_table = False
    for cells in csv.reader(io.StringIO(text)):
        cells = [(c or "").strip() for c in cells]
        if not any(cells):
            continue
        first = cells[0]
        if not in_table:
            if first.lower().startswith("tanggal"):
                in_table = True
            elif ":" in first:
                k, _, v = first.partition(":")
                meta[k.strip().lower().rstrip(".")] = v.strip()
            continue
        if len(cells) < 4:
            continue
        date = parse_date(first)
        if not date:
            # BCA marks unsettled rows with 'PEND' instead of a date
            warnings.append("Row skipped (no transaction date): %s…" % cells[1][:40])
            continue
        desc = re.sub(r"\s+", " ", cells[1]).strip()
        jumlah = cells[3].upper()
        direction = "in" if jumlah.rstrip().endswith("CR") else "out"
        amount = parse_amount(jumlah)
        balance = parse_amount(cells[4]) if len(cells) > 4 else 0.0
        if not amount:
            warnings.append("Row skipped (no amount): %s…" % desc[:40])
            continue
        # bank codes like 2505/FTSCY/WS95051 are batch codes shared by several
        # rows, so build a unique per-row fingerprint for duplicate detection
        reference = "CSV-" + hashlib.sha1(
            ("%s|%s|%s|%s" % (date, desc, amount, balance)).encode()).hexdigest()[:12].upper()
        records.append({
            "date": date, "time": "", "tx_type": desc.split("  ")[0][:60],
            "from_account": meta.get("no. rekening", meta.get("no rekening", "")),
            "va_number": "", "name": "", "merchant": "",
            "amount": amount, "transfer_type": "", "reference": reference,
            "status": "CSV", "note": "", "ok": True,
            "description": desc, "direction": direction, "balance": balance,
        })
    if not in_table:
        warnings.append("Header row 'Tanggal Transaksi' not found — is this a BCA mutasi-rekening CSV?")
    return records, warnings, meta


# ---------------------------------------------------------------------------
# BCA e-Statement PDF (REKENING GIRO / Laporan Mutasi Rekening)
# ---------------------------------------------------------------------------
# Per row: TANGGAL (DD/MM, year taken from "PERIODE : <MONTH> <YEAR>"),
# KETERANGAN (wraps over several lines), CBG, MUTASI (amount + "DB" for debit,
# plain/CR/KR for credit), and a running SALDO on some rows.

ID_MONTHS = {
    "januari": 1, "februari": 2, "maret": 3, "april": 4, "mei": 5, "juni": 6,
    "juli": 7, "agustus": 8, "september": 9, "oktober": 10, "november": 11, "desember": 12,
}
_DATE_RE = re.compile(r"^(\d{2})/(\d{2})\b")
_MONEY_RE = re.compile(r"\d{1,3}(?:,\d{3})*\.\d{2}")
_SKIP_LINE = (
    "rekening giro", "kcp ", "no. rekening", "halaman", "periode", "mata uang",
    "catatan", "apabila nasabah", "rekening ini", "telah menyetujui", "bca berhak",
    "laporan mutasi", "tanggal keterangan", "bersambung", "saldo awal :",
    "mutasi cr", "mutasi db", "saldo akhir",
)


def parse_bca_estatement_pdf(data):
    """Parse a BCA e-statement PDF. Returns (records, warnings, meta)."""
    try:
        import pdfplumber
    except ImportError:
        return [], ["pdfplumber is not installed on the server (pip install pdfplumber)."], {}

    import io as _io
    meta, lines = {}, []
    with pdfplumber.open(_io.BytesIO(data) if isinstance(data, (bytes, bytearray)) else data) as pdf:
        for page in pdf.pages:
            for raw in (page.extract_text() or "").splitlines():
                lines.append(raw.rstrip())

    year = None
    holder = None
    for ln in lines:
        low = ln.lower().strip()
        if low.startswith("periode") and ":" in ln:
            val = ln.split(":", 1)[1].strip().lower()
            for name, num in ID_MONTHS.items():
                if name in val:
                    meta["period_month"] = num
                    break
            m = re.search(r"(20\d{2})", val)
            if m:
                year = int(m.group(1))
                meta["period_year"] = year
        elif "no. rekening" in low and ":" in ln:
            meta["account"] = ln.split(":", 1)[1].strip()
            # holder name is the text before "NO. REKENING" on the same line
            before = re.split(r"(?i)no\.?\s*rekening", ln)[0].strip()
            if before and holder is None:
                holder = before
                meta["holder"] = holder
    if not year:
        return [], ["Could not read PERIODE / year from the PDF — is this a BCA e-statement?"], meta

    records, warnings = [], []
    current = None

    def flush():
        if not current:
            return
        if current["amount"] <= 0:
            return
        desc = re.sub(r"\s+", " ", " ".join(current["desc"])).strip()
        # drop the duplicate plain-number amount BCA repeats under the row
        desc = re.sub(r"\b\d{4,}\.\d{2}\b", "", desc).strip(" -")
        rec = {
            "date": current["date"], "time": "",
            "tx_type": desc.split("  ")[0][:60] if desc else "Bank transaction",
            "from_account": meta.get("account", ""), "va_number": "", "name": "", "merchant": "",
            "amount": current["amount"], "transfer_type": "",
            "reference": "PDF-" + hashlib.sha1(
                ("%s|%s|%s|%d" % (current["date"], desc, current["amount"], current["idx"]))
                .encode()).hexdigest()[:12].upper(),
            "status": "Berhasil", "note": "", "ok": True,
            "description": desc or "Bank transaction",
            "direction": current["direction"], "balance": current["balance"],
        }
        records.append(rec)

    idx = 0
    for ln in lines:
        s = ln.strip()
        if not s:
            continue
        low = s.lower()
        if any(low.startswith(p) or p in low for p in _SKIP_LINE):
            # still allow date-leading lines through (handled below); these are pure boilerplate
            if not _DATE_RE.match(s):
                continue
        m = _DATE_RE.match(s)
        if m:
            flush()
            idx += 1
            dd, mm = m.group(1), m.group(2)
            rest = s[m.end():].strip()
            if "saldo awal" in rest.lower():
                current = None
                continue
            monies = _MONEY_RE.findall(rest)
            amount, balance, direction = 0.0, 0.0, "in"
            if monies:
                amount = parse_amount(monies[0])
                # is the token straight after the amount "DB"? -> money out
                after = rest.split(monies[0], 1)[1].lstrip()
                direction = "out" if after[:2].upper() == "DB" else "in"
                if len(monies) > 1:
                    balance = parse_amount(monies[-1])
            keterangan = _MONEY_RE.sub("", rest).replace(" DB", "").replace(" CR", "").strip()
            current = {"date": "%04d-%02d-%02d" % (year, int(mm), int(dd)),
                       "amount": amount, "balance": balance, "direction": direction,
                       "desc": [keterangan] if keterangan else [], "idx": idx}
        elif current is not None:
            # continuation line of the current transaction's description
            if _MONEY_RE.fullmatch(s) or re.fullmatch(r"[\d.\-]+", s):
                continue
            current["desc"].append(s)
    flush()
    if not records:
        warnings.append("No transactions found in the PDF.")
    return records, warnings, meta


# ---------------------------------------------------------------------------
# Wallet / card platform Excel export (petty cash) — Brick-style columns
# ---------------------------------------------------------------------------
# Columns: Transaction Type, Reference ID, Status, Category, Transaction Datetime,
# Settled At, Amount, Total Fee, Currency, Description, Account Name,
# Recipient Account Number, Recipient Holder Name, Card Name, Card Holder,
# Payment Tag, Notes, GL Code, GL Name.
# Negative Amount = money out of the petty cash; positive = money in.

# transaction types that just move money between your own wallets/cards
WALLET_INTERNAL = {"INTERNAL_TRANSFER", "CARD_ADD_BALANCE", "CARD_REFUND_BALANCE"}
# suggested cost account per spending Category (codes in the standard COA)
WALLET_CATEGORY_ACCOUNT = {
    "FOOD_AND_BEVERAGE": "6900",
    "TRANSPORTATION": "6900",
    "EXPEDITION_EXPENSES": "6900",
    "OFFICE_SUPPLIES": "6600",
    "SOFTWARE": "6300",
    "TELECOMMUNICATION": "6300",
    "MISCELLANEOUS": "6900",
}


def _norm_cat(s):
    return re.sub(r"\s+", "_", str(s or "").strip().upper())


def parse_wallet_xlsx(data):
    """Parse a wallet/card transaction Excel export (petty-cash spending).
    Returns (records, warnings, meta). Money-out rows are booked as a deduction
    from the petty-cash account; internal wallet moves are flagged not-bookable.
    """
    import io as _io
    from openpyxl import load_workbook

    wb = load_workbook(_io.BytesIO(data) if isinstance(data, (bytes, bytearray)) else data, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ["The spreadsheet is empty."], {}
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {h.lower(): i for i, h in enumerate(header)}

    def col(*names):
        for n in names:
            if n.lower() in idx:
                return idx[n.lower()]
        return None

    c_type = col("Transaction Type")
    c_ref = col("Reference ID", "Reference")
    c_status = col("Status")
    c_cat = col("Category")
    c_dt = col("Transaction Datetime", "Settled At", "Date")
    c_amt = col("Amount")
    c_desc = col("Description")
    c_acct = col("Account Name")
    c_card = col("Card Name")
    c_notes = col("Notes")
    c_recip = col("Recipient Holder Name")
    if c_amt is None or c_dt is None:
        return [], ["Could not find the Amount / Transaction Datetime columns — is this the wallet export?"], {}

    records, warnings = [], []
    for n, row in enumerate(rows[1:], start=2):
        if not row or row[c_amt] in (None, "", "-"):
            continue
        status = str(row[c_status]).strip().upper() if c_status is not None else "SUCCESS"
        if status and status not in ("SUCCESS", "SETTLED", "COMPLETED"):
            continue
        try:
            amount = float(row[c_amt])
        except (TypeError, ValueError):
            continue
        if amount == 0:
            continue
        raw_dt = str(row[c_dt] or "")
        date = raw_dt[:10] if re.match(r"\d{4}-\d{2}-\d{2}", raw_dt) else parse_date(raw_dt)
        if not date:
            warnings.append("Row %d: unreadable date '%s'" % (n, raw_dt[:20]))
            continue
        ttype = str(row[c_type]).strip().upper() if c_type is not None else ""
        direction = "out" if amount < 0 else "in"
        category = _norm_cat(row[c_cat]) if c_cat is not None else ""
        desc = str(row[c_desc] or "").strip() if c_desc is not None else ""
        note = str(row[c_notes] or "").strip() if c_notes is not None and row[c_notes] not in (None, "-") else ""
        card = str(row[c_card] or "").strip() if c_card is not None and row[c_card] not in (None, "-") else ""
        recip = str(row[c_recip] or "").strip() if c_recip is not None and row[c_recip] not in (None, "-") else ""
        full_desc = " — ".join([x for x in (desc, note, recip) if x and x != "-"]) or ttype or "Wallet transaction"
        ref = str(row[c_ref] or "").strip() if c_ref is not None else ""
        internal = ttype in WALLET_INTERNAL
        suggested = None
        if direction == "out" and not internal:
            suggested = WALLET_CATEGORY_ACCOUNT.get(category, "6900")
        records.append({
            "date": date, "time": raw_dt[11:19] if len(raw_dt) > 11 else "",
            "tx_type": ttype.replace("_", " ").title(),
            "from_account": str(row[c_acct] or "").strip() if c_acct is not None else "",
            "va_number": card, "name": recip, "merchant": recip,
            "amount": round(abs(amount), 2), "transfer_type": "",
            "reference": ref or ("WLT-" + hashlib.sha1(
                ("%s|%s|%s|%d" % (date, full_desc, amount, n)).encode()).hexdigest()[:12].upper()),
            "status": status.title(), "note": note, "ok": True,
            "description": full_desc, "direction": direction,
            "category": category, "internal": internal, "suggested_code": suggested,
            "balance": 0.0,
        })
    if not records:
        warnings.append("No usable transactions found in the spreadsheet.")
    meta = {"account": str(rows[1][c_acct]) if c_acct is not None and len(rows) > 1 else "",
            "rows": len(records)}
    return records, warnings, meta
